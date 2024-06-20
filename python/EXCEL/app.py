import openpyxl
import openpyxl.utils
import os.path

#[0]
# TODO products 리스트의 데이터를 엑셀에 저장하시오
# An : 제품명, Bn : 보유수량, Cn : 제품단가, Dn : 단가총액(=Bn*Cn)
# 마지막 라인에는 An : '총액', Dn : 제품보유단가총액(=SUM(D1:Dn))
products = [
{'name': 'CPU', 'count': 100, 'price': 350000},
{'name': 'BOARD', 'count': 120, 'price': 125000},
{'name': 'RAM', 'count': 98, 'price': 60000},
{'name': 'SSD', 'count': 85, 'price': 95000},
{'name': 'HDD', 'count': 112, 'price': 80000},
{'name': 'PSU', 'count': 105, 'price': 89000},
{'name': 'MONITOR', 'count': 80, 'price': 280000},
{'name': 'KEYBOARD', 'count': 90, 'price': 23000},
{'name': 'MOUSE', 'count': 110, 'price': 25000}
]

headers = ['제품명', '보유수량', '제품단가', '단가총액']
excelFile = os.path.join('.', 'data.xlsx')
wb = openpyxl.load_workbook(excelFile)
ws = wb.active
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = '재고보유현황'
ws.append(headers)

row = 2

for tp in products:
    ws.append([value for key, value in tp.items()])
    ws.cell(row=row, column=4, value=f'=B{row}*C{row}')
    row += 1
ws.append(['총액', '', '', f'=SUM(D2:D{row-1})'])

wb.save(excelFile)
wb.close()

'''[8]
excelFile = os.path.join('.', 'data.xlsx')
wb = openpyxl.load_workbook(excelFile)
ws = wb.active
products = [
    {'NAME' : 'CPU', 'PRICE' : 250000, 'STOCK' : 10},
    {'NAME' : 'CPU', 'PRICE' : 100000, 'STOCK' : 20},
    {'NAME' : 'CPU', 'PRICE' : 120000, 'STOCK' : 15}
]
row = 1
for item in products:
    #ws.append([value for key, value in item.items()])
    A = openpyxl.utils.get_column_letter(2)
    B = openpyxl.utils.get_column_letter(3)
    ws.cell(row=row, column='D', value='={A}{row} * {B}{row}')
    row += 1
wb.save(excelFile)
wb.close()
'''

'''[7]
for product in products:
    for key, value in product.items():
        ws.append([key, value])
wb.save(excelFile)
wb.close()
'''

'''[6]
products = { 'cpu' : 25, 'ram' : 20, 'ssd' : 22 }
row = 1
for key, value in products.items():
    # ws.append([key, value])
    ws.cell(row=row, column=1, value=key)
    ws.cell(row=row, column=2, value=key)
    row += 1
wb.save(excelFile)
wb.close()
'''

'''[5]
ws['A1'] = 10
ws['B1'] = 20
ws['C1'] = '=A1 + B1'
wb.save(excelFile)
wb.close()
'''

'''[4]
status = [10, 20, 30, 40]
ws = wb['IP주소별 접속통계']
ws['A1'] = '이름'       # excel파일 A1 칸에 10 입력
ws.cell(row=1, column=2, value='=SUM(A1:A3)')        # 해당 행,열에 value(값 or 엑셀함수가능)
ws.append(status)       # statue 리스트 요소들을 추가
wb.save(excelFile)
wb.close()
'''

'''[3]
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'IP주소별 접속통계'
wb.save('data.xlsx')
wb.close()
'''

'''[2]
print(dir())
print(dir(openpyxl))
print(dir(openpyxl.Workbook))
'''

'''[1]
import mymodule

def main():
    pass

if __name__ == '__name__':
    main()
    print(__name__)
'''